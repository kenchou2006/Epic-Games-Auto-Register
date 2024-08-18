from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import string
import random
import openpyxl
import os.path
import imaplib
import email
import re
import time

#在此變更帳號
mail_account = "test.account"
mail_server = "@gmail.com"
imap_password = 'mgksdnvrmqgajmtz'

def generate_password():
    length = random.randint(12, 17)
    chars = string.ascii_lowercase + string.digits
    password = ''.join(random.choice(chars) for i in range(length))
    return password
def generate_displayname():
    length = random.randint(8, 12)
    chars = string.ascii_lowercase + string.digits
    displayname = ''.join(random.choice(chars) for i in range(length))
    return displayname
filename = 'account.xlsx'
if not os.path.isfile(filename):
    wb = openpyxl.Workbook()
    wb.active.title = 'Account'
    sheet = wb.active
    sheet['A1'] = 'Email'
    sheet['B1'] = 'Password'
    sheet['C1'] = 'DisplayName'
    sheet['D1'] = 'Used By'
    last_row = 1
    print(f"The Excel file {filename} has been created.")
else:
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    last_row = sheet.max_row
    print(f"The Excel file {filename} has been updated. The total rows used: {last_row}.")
mail_gea = "+"
reg_mail = mail_account + mail_gea + str(last_row) + mail_server
data = {'email': reg_mail, 'password': generate_password(), 'displayname': generate_displayname()}
if last_row > 1 and sheet.cell(row=last_row, column=1).value is not None:
    sheet.append(['', ''])
sheet.cell(row=last_row+1, column=1).value = data['email']
sheet.cell(row=last_row+1, column=2).value = data['password']
sheet.cell(row=last_row+1, column=3).value = data['displayname']
wb.save(filename)
wb.close()
print(data['email'])
print(data['password'])
print(data['displayname'])
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.epicgames.com/id/register/date-of-birth")
time.sleep(3)
Data_of_Birth_Year = driver.find_element_by_xpath('//*[@id="year"]')
Data_of_Birth_Year.click()
Data_of_Birth_Year.send_keys("2000")
time.sleep(3)
Data_of_Birth_Month = driver.find_element_by_xpath('//*[@id="month"]')
Data_of_Birth_Month.click()
Data_of_Birth_Month_Jan = driver.find_element_by_xpath('//*[@id="menu-"]/div[3]/ul/li[1]')
Data_of_Birth_Month_Jan.click()
time.sleep(3)
Data_of_Birth_Day = driver.find_element_by_xpath('//*[@id="day"]')
Data_of_Birth_Day.click()
Data_of_Birth_Day1 = driver.find_element_by_xpath('//*[@id="menu-"]/div[3]/ul/li[1]')
Data_of_Birth_Day1.click()
time.sleep(3)
Data_of_Birth_Continue = driver.find_element_by_xpath('//*[@id="continue"]')
Data_of_Birth_Continue.click()
time.sleep(3)
Sing_Up_First_Name = driver.find_element_by_xpath('//*[@id="name"]')
Sing_Up_First_Name.click()
Sing_Up_First_Name.send_keys("咿")
time.sleep(3)
Sing_Up_Last_Name = driver.find_element_by_xpath('//*[@id="lastName"]')
Sing_Up_Last_Name.click()
Sing_Up_Last_Name.send_keys("二三")
time.sleep(3)
Sing_Up_Display_Name = driver.find_element_by_xpath('//*[@id="displayName"]')
Sing_Up_Display_Name.click()
Sing_Up_Display_Name.send_keys(data['displayname'])
time.sleep(3)
Sing_Up_Email = driver.find_element_by_xpath('//*[@id="email"]')
Sing_Up_Email.click()
Sing_Up_Email.send_keys(data['email'])
time.sleep(3)
Sing_Up_Password = driver.find_element_by_xpath('//*[@id="password"]')
Sing_Up_Password.click()
Sing_Up_Password.send_keys(data['password'])
time.sleep(3)
Sing_Up_Agree = driver.find_element_by_xpath('//*[@id="tos"]')
Sing_Up_Agree.click()
time.sleep(3)
Sing_Up_Continue = driver.find_element_by_xpath('//*[@id="btn-submit"]')
Sing_Up_Continue.click()
time.sleep(3)
imap_address = mail_account+mail_server
imap_server = imaplib.IMAP4_SSL('imap.gmail.com')
imap_server.login(imap_address, imap_password)
imap_server.select('inbox')
status, email_ids = imap_server.search(None, '(UNSEEN SUBJECT "Epic")')
email_ids = email_ids[0].split()
ignore_list = ['313131', '202020', '858585']
for email_id in email_ids:
    status, email_data = imap_server.fetch(email_id, '(RFC822)')
    email_message = email.message_from_bytes(email_data[0][1])
    sender = email.utils.parseaddr(email_message['From'])[1]
    subject = email_message['Subject']
    body = ''
    if email_message.is_multipart():
        for part in email_message.walk():
            content_type = part.get_content_type()
            if content_type == 'text/plain':
                body = part.get_payload(decode=True).decode('utf-8')
                break
            elif content_type == 'text/html':
                body = part.get_payload(decode=True).decode('utf-8')
                break
    else:
        body = email_message.get_payload(decode=True).decode('utf-8')
    codes = re.findall(r'\b\d{6}\b', body)
    for code in codes:
        if code in ignore_list:
            continue
        print("IMAP Into")
        print("From: ",sender)
        print("Subject: ",subject)
        print("Code: ",code)
        digits = [int(d) for d in str(code)]
        digit1, digit2, digit3, digit4, digit5, digit6 = digits
imap_server.close()
imap_server.logout()
time.sleep(3)
Verify_Number_1 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[1]/div/input')
Verify_Number_1.click()
Verify_Number_1.send_keys(digit1)
time.sleep(3)
Verify_Number_2 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[2]/div/input')
Verify_Number_2.click()
Verify_Number_2.send_keys(digit2)
time.sleep(3)
Verify_Number_3 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[3]/div/input')
Verify_Number_3.click()
Verify_Number_3.send_keys(digit3)
time.sleep(3)
Verify_Number_4 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[4]/div/input')
Verify_Number_4.click()
Verify_Number_4.send_keys(digit4)
time.sleep(3)
Verify_Number_5 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[5]/div/input')
Verify_Number_5.click()
Verify_Number_5.send_keys(digit5)
time.sleep(3)
Verify_Number_6 = driver.find_element_by_xpath('//*[@id="modal-content"]/div[2]/div/div[1]/form/div[1]/div/div[6]/div/input')
Verify_Number_6.click()
Verify_Number_6.send_keys(digit6)
time.sleep(3)
Verify_Buttom = driver.find_element_by_xpath('//*[@id="continue"]/span')
Verify_Buttom.click()
time.sleep(3)
driver.get("https://store.epicgames.com/logout")
#driver.quit()
